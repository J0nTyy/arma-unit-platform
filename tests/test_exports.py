"""ExportService: file naming, overwrite protection, Excel compatibility."""

from openpyxl import load_workbook

from app.services.exports import ExportService

HEADERS = ["Name", "Count"]
ROWS = [["Alpha", 1], ["Bravo, Charlie", 2]]  # comma forces CSV quoting


def test_dated_export_names_and_collision_suffix(tmp_path):
    service = ExportService()
    first = service.dated_export(tmp_path, "members", HEADERS, ROWS)
    second = service.dated_export(tmp_path, "members", HEADERS, ROWS)
    third = service.dated_export(tmp_path, "members", HEADERS, ROWS)

    assert first[0].name.startswith("members_20") and first[0].suffix == ".csv"
    assert second[0].stem.endswith("_2")  # never silently overwritten
    assert third[0].stem.endswith("_3")
    assert len(list(tmp_path.glob("members_*.csv"))) == 3


def test_csv_is_excel_friendly_utf8_bom(tmp_path):
    service = ExportService()
    (path,) = service.dated_export(tmp_path, "members", HEADERS, ROWS)
    raw = path.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM for Excel double-click
    text = raw.decode("utf-8-sig")
    assert text.splitlines()[0] == "Name,Count"
    assert '"Bravo, Charlie"' in text  # proper quoting


def test_xlsx_roundtrip(tmp_path):
    service = ExportService()
    paths = service.dated_export(tmp_path, "ops", HEADERS, ROWS, formats=("csv", "xlsx"))
    xlsx = next(p for p in paths if p.suffix == ".xlsx")
    sheet = load_workbook(xlsx).active
    assert [c.value for c in sheet[1]] == HEADERS
    assert sheet["A2"].value == "Alpha" and sheet["B3"].value == 2
    assert sheet.freeze_panes == "A2"


def test_snapshot_overwrites_in_place(tmp_path):
    service = ExportService()
    path = service.snapshot(tmp_path, "members", HEADERS, [["Old", 1]])
    again = service.snapshot(tmp_path, "members", HEADERS, [["New", 2]])
    assert path == again
    assert "New" in path.read_text(encoding="utf-8-sig")
    assert len(list(tmp_path.glob("members*.csv"))) == 1  # no _2 files


def test_dated_workbook_multi_sheet_and_collision(tmp_path):
    service = ExportService()
    sheets = {"Members": (HEADERS, ROWS), "Ops": (HEADERS, [])}
    first = service.dated_workbook(tmp_path, "unit-data", sheets)
    second = service.dated_workbook(tmp_path, "unit-data", sheets)

    workbook = load_workbook(first)
    assert workbook.sheetnames == ["Members", "Ops"]
    assert workbook["Members"]["A2"].value == "Alpha"
    assert workbook["Ops"]["A1"].value == "Name"  # empty sheet keeps headers
    assert second.stem.endswith("_2")  # same-day rerun never overwrites


def test_prune_dated_keeps_newest(tmp_path):
    service = ExportService()
    sheets = {"Members": (HEADERS, ROWS)}
    # Creation order lives in the filenames (date + _N counter), so even
    # same-instant writes prune correctly.
    created = [service.dated_workbook(tmp_path, "unit-data", sheets) for _ in range(4)]

    removed = ExportService.prune_dated(tmp_path, "unit-data", keep=2)
    assert [p.name for p in removed] == [p.name for p in created[:2]]
    survivors = {p.name for p in tmp_path.glob("unit-data_*.xlsx")}
    assert survivors == {p.name for p in created[2:]}


def test_empty_dataset_still_writes_headers(tmp_path):
    service = ExportService()
    paths = ExportService().dated_export(tmp_path, "empty", HEADERS, [], formats=("csv", "xlsx"))
    assert paths[0].read_text(encoding="utf-8-sig").strip() == "Name,Count"
    sheet = load_workbook(paths[1]).active
    assert [c.value for c in sheet[1]] == HEADERS
