"""Data exports: dataset building from real services, files on disk."""

import csv
from datetime import datetime, timedelta, timezone

from app.database.models.operation import AttendanceStatus, OperationStatus
from app.database.models.player import FinalAttendance
from app.services import (
    AttendanceService,
    DataExportService,
    MemoryService,
    OperationService,
    PlayerService,
)


async def _scenario(database):
    """One member signs up, attends, gets a cert — the full data trail."""
    players = PlayerService(database)
    operations = OperationService(database)
    attendance = AttendanceService(database)

    await players.update_preferences(
        1, 100, "Kartikey", primary_role="medic", steam_id="76561198000000001"
    )
    await players.grant_qualification(1, 100, "medic", granted_by=999)

    operation = await operations.create_operation(
        guild_id=1, mission_id="OP-001", mission_name="Blackout",
        mission_status="ready",
        scheduled_at_utc=datetime.now(timezone.utc) + timedelta(hours=1),
        tz_name="UTC", created_by=1,
    )
    await operations.mark_published(operation.id, channel_id=1, message_id=2)
    await operations.set_attendance(operation.id, 100, "Kartikey", AttendanceStatus.ATTENDING)
    await operations.transition(operation.id, OperationStatus.ACTIVE)
    await operations.transition(operation.id, OperationStatus.COMPLETED)
    await attendance.set_final_status(
        operation.id, 1, 100, "Kartikey", FinalAttendance.ATTENDED, changed_by=999
    )


def _read_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    return rows[0], rows[1:]


async def test_datasets_cover_all_tables(database):
    await _scenario(database)
    datasets = await DataExportService(database).datasets(1)
    assert set(datasets) == {
        "members", "operations", "attendance", "certifications", "missions"
    }

    headers, rows = datasets["members"]
    assert headers[0] == "Name"
    member = rows[0]
    assert member[0] == "Kartikey"
    assert "Combat Medic" in member[10]
    assert member[11] == 1  # attended
    assert member[14] == "100%"

    _, operation_rows = datasets["operations"]
    assert operation_rows[0][1] == "Blackout"
    assert operation_rows[0][6] == 1  # signed up
    assert operation_rows[0][7] == 1  # attended


async def test_attendance_is_one_row_per_record(database):
    await _scenario(database)
    headers, rows = (await DataExportService(database).datasets(1))["attendance"]
    # The spec's table shape: filterable rows, never per-player op columns.
    assert headers[:8] == [
        "Player", "Operation", "Mission", "Operation date", "Signup",
        "Final status", "Role", "Notes",
    ]
    (row,) = rows
    assert row[0] == "Kartikey"
    assert row[1] == "Blackout"
    assert row[4] == "attending"   # what they said beforehand
    assert row[5] == "attended"    # what staff finalized
    assert row[6] == "Medic"       # role preference label


async def test_export_is_guild_scoped(database):
    players = PlayerService(database)
    await players.get_or_create(1, 100, "InGuild")
    await players.get_or_create(2, 200, "OtherGuild")
    _, rows = (await DataExportService(database).datasets(1))["members"]
    assert [row[0] for row in rows] == ["InGuild"]


async def test_export_workbook_is_one_file_with_all_sheets(database, tmp_path):
    from openpyxl import load_workbook

    await _scenario(database)
    service = DataExportService(database)
    exports = tmp_path / "exports"

    path, counts = await service.export_workbook(1, exports)
    assert path.name.startswith("unit-data_") and path.suffix == ".xlsx"
    assert counts["members"] == 1 and counts["attendance"] == 1

    workbook = load_workbook(path)
    assert set(workbook.sheetnames) == {
        "Members", "Operations", "Attendance", "Certifications", "Missions"
    }
    attendance = workbook["Attendance"]
    assert attendance["A1"].value == "Player" and attendance["A2"].value == "Kartikey"

    # The always-current CSVs are refreshed by the same export call.
    latest = exports / "latest"
    assert sorted(p.name for p in latest.glob("*.csv")) == [
        "attendance.csv", "certifications.csv", "members.csv",
        "missions.csv", "operations.csv",
    ]
    # Snapshots regenerate in place — running again adds no extra files there.
    await service.write_snapshots(1, exports)
    assert len(list(latest.glob("*.csv"))) == 5


async def test_export_workbook_retention_caps_file_count(database, tmp_path):
    await _scenario(database)
    service = DataExportService(database)
    exports = tmp_path / "exports"

    paths = []
    for _ in range(13):
        path, _counts = await service.export_workbook(1, exports)
        paths.append(path)

    remaining = sorted(p.name for p in exports.glob("unit-data_*.xlsx"))
    assert len(remaining) == 10  # never more than the newest 10
    assert paths[-1].name in remaining      # newest kept
    assert paths[0].name not in remaining   # oldest pruned


async def test_memory_snapshot_is_readable_markdown(database, tmp_path):
    memories = MemoryService(database)
    await memories.remember(1, "Op nights moved to Fridays at 2000", author_id=7)
    await memories.remember(
        1, "Server maintenance this weekend", author_id=7, days_valid=3
    )
    await memories.remember(2, "Other guild secret", author_id=9)

    path = await DataExportService(database).write_memory_snapshot(1, tmp_path / "memory")
    text = path.read_text(encoding="utf-8")
    assert path.name == "memories.md"
    assert "Fridays at 2000" in text
    assert "expires:" in text            # temporary fact is marked
    assert "Other guild secret" not in text  # guild isolation
    assert "canonical" in text           # documents that DB is the source
