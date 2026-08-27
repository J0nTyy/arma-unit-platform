"""Tabular file exports (CSV / XLSX).

ExportService knows how to write tables to files; it knows nothing about
what the tables contain (datasets live in DataExportService). Two modes,
matching the canonical-vs-generated rule in data/README.md:

- **Dated exports** — staff-triggered, written once, never silently
  overwritten (collisions get ``_2``, ``_3``…). ``dated_workbook`` packs
  several datasets into ONE Excel file with one sheet per dataset, and
  ``prune_dated`` caps how many old exports pile up.
- **Snapshots** — regularly regenerated "latest state" files that are
  overwritten in place: ``exports/latest/members.csv``.

CSV files are written as UTF-8 with BOM so Excel opens them correctly
without an import wizard; XLSX via openpyxl.
"""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

Rows = Sequence[Sequence[Any]]


class ExportService:
    def dated_export(
        self,
        directory: Path,
        name: str,
        headers: Sequence[str],
        rows: Rows,
        *,
        formats: Sequence[str] = ("csv",),
    ) -> list[Path]:
        """Write a dated, never-overwritten export. Returns created paths."""
        directory.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        written: list[Path] = []
        for fmt in formats:
            path = self._free_path(directory, f"{name}_{stamp}", fmt)
            self._write(path, fmt, headers, rows)
            written.append(path)
        return written

    def dated_workbook(
        self,
        directory: Path,
        name: str,
        sheets: dict[str, tuple[Sequence[str], Rows]],
    ) -> Path:
        """One dated Excel workbook, one sheet per dataset. Never overwrites."""
        directory.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        path = self._free_path(directory, f"{name}_{stamp}", "xlsx")
        workbook = Workbook()
        workbook.remove(workbook.active)
        for title, (headers, rows) in sheets.items():
            sheet = workbook.create_sheet(title=title[:31])
            self._fill_sheet(sheet, headers, rows)
        workbook.save(path)
        return path

    def snapshot(
        self, directory: Path, name: str, headers: Sequence[str], rows: Rows
    ) -> Path:
        """Write/overwrite a 'latest state' CSV snapshot (generated file)."""
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / f"{name}.csv"
        self._write(path, "csv", headers, rows)
        return path

    @staticmethod
    def prune_dated(directory: Path, name: str, keep: int) -> list[Path]:
        """Delete the oldest dated exports beyond `keep`. Returns deletions.

        Creation order is read from the filename itself — the date plus the
        collision counter (`_2`, `_3`…) — because same-day files can share a
        filesystem timestamp.
        """
        if keep <= 0 or not directory.is_dir():
            return []

        def created_order(path: Path) -> tuple[str, int]:
            stem = path.stem.removeprefix(f"{name}_")  # "2026-08-27" or "2026-08-27_2"
            date_part, _, counter = stem.partition("_")
            return (date_part, int(counter) if counter.isdigit() else 1)

        files = sorted(directory.glob(f"{name}_*.xlsx"), key=created_order)
        removed = files[:-keep] if len(files) > keep else []
        for file in removed:
            file.unlink()
        return removed

    @staticmethod
    def _free_path(directory: Path, stem: str, fmt: str) -> Path:
        """Monotonic naming: stem.fmt, then stem_2.fmt, stem_3.fmt, ...

        Counters are never reused, even after older files are deleted by
        pruning — otherwise a new file could take over a freed-up early name
        and later be mistaken for the oldest. Name order == creation order.
        """
        highest = 0
        for existing in directory.glob(f"{stem}*.{fmt}"):
            remainder = existing.stem.removeprefix(stem)  # "" or "_<n>"
            if remainder == "":
                highest = max(highest, 1)
            elif remainder.startswith("_") and remainder[1:].isdigit():
                highest = max(highest, int(remainder[1:]))
        if highest == 0:
            return directory / f"{stem}.{fmt}"
        return directory / f"{stem}_{highest + 1}.{fmt}"

    @staticmethod
    def _fill_sheet(sheet, headers: Sequence[str], rows: Rows) -> None:
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for row in rows:
            sheet.append(list(row))
        for index, header in enumerate(headers, start=1):
            width = max(
                len(str(header)),
                *(len(str(row[index - 1])) for row in rows) if rows else (0,),
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 40)

    @classmethod
    def _write(cls, path: Path, fmt: str, headers: Sequence[str], rows: Rows) -> None:
        if fmt == "csv":
            # utf-8-sig: Excel needs the BOM to detect UTF-8 on double-click.
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                writer.writerows(rows)
        elif fmt == "xlsx":
            workbook = Workbook()
            cls._fill_sheet(workbook.active, headers, rows)
            workbook.save(path)
        else:
            raise ValueError(f"unsupported export format: {fmt}")
